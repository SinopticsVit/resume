# -*- coding: utf-8 -*-
"""Генератор Excel-бюджета представительства «Детского Мира» в КНР (6 месяцев).

Источник истины: исходные цифры на листе "Зарплаты и найм".
Сводные листы ("Сводка", "OPEX помесячно", "Денежный поток") считаются
формулами от исходных данных.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MONTHS = ["M1", "M2", "M3", "M4", "M5", "M6"]

# Значения по умолчанию (как в отредактированном v2)
RUB_RATE_DEFAULT = 12
LOCAL_SOCIAL_RATE_DEFAULT = 0.35
EXPAT_SOCIAL_RATE_DEFAULT = 0.15
RO_TAX_RATE_DEFAULT = 0.10

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F3864")
H_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD = Font(name="Calibri", size=11, bold=True)
NORMAL = Font(name="Calibri", size=11)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="808080")
H_FILL = PatternFill("solid", fgColor="1F3864")
SUBTOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
CAPEX_FILL = PatternFill("solid", fgColor="E2EFDA")
SOURCE_FILL = PatternFill("solid", fgColor="F2F2F2")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

NUMFMT = "#,##0"
NUMFMT_RATE = "0.00%"


def style_header(cell):
    cell.font = H_FONT
    cell.fill = H_FILL
    cell.alignment = CENTER
    cell.border = BORDER


def num(cell, value, fmt=NUMFMT, bold=False, fill=None):
    if value is not None:
        cell.value = value
    cell.number_format = fmt
    cell.font = BOLD if bold else NORMAL
    cell.alignment = RIGHT
    cell.border = BORDER
    if fill:
        cell.fill = fill


def txt(cell, value, bold=False, fill=None, align=LEFT):
    cell.value = value
    cell.font = BOLD if bold else NORMAL
    cell.alignment = align
    cell.border = BORDER
    if fill:
        cell.fill = fill


salary_rows = [
    # role, type, base, compensation, relocation_once, M1..M6
    ("Директор (RU, экспат)", "Экспат", 50000, 20000, 0, ["●", "●", "●", "●", "●", "●"]),
    ("Технолог (RU, экспат)", "Экспат", 30000, 20000, 20000, ["", "", "", "", "", "●"]),
    ("Кат. менеджер — игрушки (CN)", "Локальный", 18500, 0, 0, ["", "●", "●", "●", "●", "●"]),
    ("Кат. менеджер — одежда (CN)", "Локальный", 18500, 0, 0, ["", "", "", "●", "●", "●"]),
    ("Кат. менеджер — аксессуары (CN)", "Локальный", 18500, 0, 0, ["", "", "", "", "", "●"]),
    ("Руководитель ОТК (CN)", "Локальный", 18000, 0, 0, ["", "●", "●", "●", "●", "●"]),
    ("Бухгалтер штатный (CN)", "Локальный", 16000, 0, 0, ["", "", "●", "●", "●", "●"]),
    ("Инспектор ОТК №1 (CN)", "Локальный", 10000, 0, 0, ["", "", "", "●", "●", "●"]),
    ("Инспектор ОТК №2 (CN)", "Локальный", 10000, 0, 0, ["", "", "", "", "", "●"]),
    ("Офис-администратор (CN)", "Локальный", 10000, 0, 0, ["", "", "●", "●", "●", "●"]),
]

capex_items = [
    ("Юридическая регистрация RO", 20000, "6–10 недель"),
    ("Регистрация ТМ в CNIPA (3 марки)", 10000, "до начала sourcing"),
    ("Офис: депозит + ремонт/мебель (~100 м²)", 15000, "Шанхай Grade A/B"),
    ("IT-капекс (ноутбуки, VPN, ERP-коннектор)", 50000, "подключение к HQ"),
    ("Стартовые командировки", 86000, "3–4 поездки"),
    ("Оборудование ОТК", 10000, "полевой набор"),
    ("Резерв (contingency ~10%)", 48000, ""),
]

# Распределение CAPEX по месяцам (сумма = 239 000 CNY, совпадает с CAPEX-таблицей).
capex_by_month = [144000, 72000, 13000, 10000, 0, 0]

other_opex_assumptions = {
    "Аутсорс-бухгалтерия": [8000, 8000, 8000, 0, 0, 0],
    "Аренда офиса": [6000, 6000, 20000, 20000, 20000, 20000],
    "Командировки и QC-выезды": [5000, 10000, 20000, 30000, 35000, 35000],
    "IT / связь (opex)": [3000, 5000, 5000, 5000, 5000, 5000],
    "Юр./комплаенс-ретейнер": [10000, 8000, 5000, 5000, 5000, 5000],
    "Прочее (перевод, ресёрч, hospit.)": [3000, 3000, 3000, 3000, 3000, 3000],
}

salary_sheet_name = "Зарплаты и найм"
salary_start_row = 5
salary_end_row = salary_start_row + len(salary_rows) - 1
month_cols_salary = ["G", "H", "I", "J", "K", "L"]

rate_local_cell = "D17"
rate_expat_cell = "D18"
rate_tax_cell = "D19"
rate_rub_cell = "D20"

other_opex_start_row = 24
capex_month_row = 32
capex_items_start_row = 35
capex_total_row = 4 + len(capex_items) + 1
opex_total_row = 15


def salary_formula(month_idx: int) -> str:
    """ФОТ месяц = активные (база+компенсации) + релокация в первый активный месяц."""
    cur_col = month_cols_salary[month_idx]
    active = f"'{salary_sheet_name}'!${cur_col}${salary_start_row}:${cur_col}${salary_end_row}"
    base = f"'{salary_sheet_name}'!$D${salary_start_row}:$D${salary_end_row}"
    comp = f"'{salary_sheet_name}'!$E${salary_start_row}:$E${salary_end_row}"
    reloc = f"'{salary_sheet_name}'!$F${salary_start_row}:$F${salary_end_row}"

    recurring = f'SUMPRODUCT(--({active}="●"),{base}+{comp})'
    if month_idx == 0:
        one_time = f'SUMPRODUCT(--({active}="●"),{reloc})'
    else:
        prev_col = month_cols_salary[month_idx - 1]
        prev_active = f"'{salary_sheet_name}'!${prev_col}${salary_start_row}:${prev_col}${salary_end_row}"
        one_time = f'SUMPRODUCT(--({active}="●"),--({prev_active}<>"●"),{reloc})'
    return f"={recurring}+{one_time}"


def social_formula(month_idx: int) -> str:
    cur_col = month_cols_salary[month_idx]
    active = f"'{salary_sheet_name}'!${cur_col}${salary_start_row}:${cur_col}${salary_end_row}"
    role_type = f"'{salary_sheet_name}'!$C${salary_start_row}:$C${salary_end_row}"
    base_comp = (
        f"'{salary_sheet_name}'!$D${salary_start_row}:$D${salary_end_row}"
        f"+'{salary_sheet_name}'!$E${salary_start_row}:$E${salary_end_row}"
    )
    local_part = (
        f'SUMPRODUCT(--({active}="●"),--({role_type}="Локальный"),{base_comp})'
        f"*'{salary_sheet_name}'!${rate_local_cell}"
    )
    expat_part = (
        f'SUMPRODUCT(--({active}="●"),--({role_type}="Экспат"),{base_comp})'
        f"*'{salary_sheet_name}'!${rate_expat_cell}"
    )
    return f"=ROUND({local_part}+{expat_part},0)"


wb = Workbook()

# === Сводка ===
ws = wb.active
ws.title = "Сводка"
ws.sheet_view.showGridLines = False
for c, w in [("A", 4), ("B", 42), ("C", 18), ("D", 18)]:
    ws.column_dimensions[c].width = w
ws["B2"] = "Бюджет представительства «Детского Мира» в КНР"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Помесячные сметы на 6 месяцев (формулы от листа «Зарплаты и найм»)"
ws["B3"].font = NOTE_FONT
for c, label in zip("BCD", ["Блок", "CNY", "RUB"]):
    style_header(ws[f"{c}6"])
    ws[f"{c}6"] = label

txt(ws["B7"], "CAPEX (единовременно, старт)", fill=CAPEX_FILL)
ws["C7"] = f"=CAPEX!C{capex_total_row}"
num(ws["C7"], None, fill=CAPEX_FILL)
ws["D7"] = f"=C7*'{salary_sheet_name}'!${rate_rub_cell}"
num(ws["D7"], None, fill=CAPEX_FILL)

txt(ws["B8"], "OPEX (6 мес., вкл. налог RO)", fill=SUBTOTAL_FILL)
ws["C8"] = f"='OPEX помесячно'!I{opex_total_row}"
num(ws["C8"], None, fill=SUBTOTAL_FILL)
ws["D8"] = f"=C8*'{salary_sheet_name}'!${rate_rub_cell}"
num(ws["D8"], None, fill=SUBTOTAL_FILL)

txt(ws["B9"], "ИТОГО за 6 месяцев", bold=True, fill=TOTAL_FILL)
ws["C9"] = "=C7+C8"
num(ws["C9"], None, bold=True, fill=TOTAL_FILL)
ws["D9"] = f"=C9*'{salary_sheet_name}'!${rate_rub_cell}"
num(ws["D9"], None, bold=True, fill=TOTAL_FILL)

# === CAPEX ===
ws = wb.create_sheet("CAPEX")
ws.sheet_view.showGridLines = False
for c, w in [("A", 4), ("B", 44), ("C", 16), ("D", 16), ("E", 26)]:
    ws.column_dimensions[c].width = w
ws["B2"] = "CAPEX — единовременные затраты старта"
ws["B2"].font = TITLE_FONT
for c, label in zip("BCDE", ["Статья", "CNY", "RUB", "Комментарий"]):
    style_header(ws[f"{c}4"])
    ws[f"{c}4"] = label

row = 5
for idx, (name, _, comment) in enumerate(capex_items):
    src_row = capex_items_start_row + idx
    txt(ws[f"B{row}"], name)
    ws[f"C{row}"] = f"='{salary_sheet_name}'!C{src_row}"
    num(ws[f"C{row}"], None)
    ws[f"D{row}"] = f"=C{row}*'{salary_sheet_name}'!${rate_rub_cell}"
    num(ws[f"D{row}"], None)
    txt(ws[f"E{row}"], comment)
    row += 1
txt(ws[f"B{row}"], "Итого CAPEX", bold=True, fill=TOTAL_FILL)
ws[f"C{row}"] = f"=SUM(C5:C{row-1})"
num(ws[f"C{row}"], None, bold=True, fill=TOTAL_FILL)
ws[f"D{row}"] = f"=C{row}*'{salary_sheet_name}'!${rate_rub_cell}"
num(ws[f"D{row}"], None, bold=True, fill=TOTAL_FILL)
txt(ws[f"E{row}"], "", fill=TOTAL_FILL)

# === Зарплаты и найм (источник) ===
ws = wb.create_sheet(salary_sheet_name)
ws.sheet_view.showGridLines = False
for c, w in [("A", 3), ("B", 34), ("C", 13), ("D", 17), ("E", 23), ("F", 18)]:
    ws.column_dimensions[c].width = w
for col in "GHIJKL":
    ws.column_dimensions[col].width = 7
ws["B2"] = "Исходные данные бюджета (редактировать серые ячейки)"
ws["B2"].font = TITLE_FONT
headers = [
    "Роль", "Тип", "CNY/мес\nЗарплата+налоги",
    "Компенсации жилье + обучение детей", "Релокация единоразово",
] + MONTHS
for i, label in enumerate(headers):
    col = get_column_letter(2 + i)
    style_header(ws[f"{col}4"])
    ws[f"{col}4"] = label

row = salary_start_row
for role, rtype, base, comp, reloc, sched in salary_rows:
    txt(ws[f"B{row}"], role)
    txt(ws[f"C{row}"], rtype, align=CENTER)
    num(ws[f"D{row}"], base, fill=SOURCE_FILL)
    num(ws[f"E{row}"], comp, fill=SOURCE_FILL)
    num(ws[f"F{row}"], reloc, fill=SOURCE_FILL)
    for i, mark in enumerate(sched):
        col = month_cols_salary[i]
        txt(ws[f"{col}{row}"], mark, align=CENTER, fill=SOURCE_FILL)
    row += 1
txt(ws[f"B{row}"], "Активных сотрудников", bold=True, fill=SUBTOTAL_FILL)
for col in month_cols_salary:
    ws[f"{col}{row}"] = f'=COUNTIF({col}{salary_start_row}:{col}{salary_end_row},"●")'
    num(ws[f"{col}{row}"], None, bold=True, fill=SUBTOTAL_FILL)
    ws[f"{col}{row}"].alignment = CENTER

for r, label in [(17, "Ставка соцвзносов (локальные)"), (18, "Ставка соцвзносов (экспаты)"),
                 (19, "Налог RO (deemed)"), (20, "Курс CNY->RUB")]:
    txt(ws[f"B{r}"], label)
num(ws[rate_local_cell], LOCAL_SOCIAL_RATE_DEFAULT, fmt=NUMFMT_RATE, fill=SOURCE_FILL)
num(ws[rate_expat_cell], EXPAT_SOCIAL_RATE_DEFAULT, fmt=NUMFMT_RATE, fill=SOURCE_FILL)
num(ws[rate_tax_cell], RO_TAX_RATE_DEFAULT, fmt=NUMFMT_RATE, fill=SOURCE_FILL)
num(ws[rate_rub_cell], RUB_RATE_DEFAULT, fill=SOURCE_FILL)

ws["B22"] = "Прочие OPEX-драйверы"
ws["B22"].font = BOLD
for i, label in enumerate(["Статья"] + MONTHS):
    col = get_column_letter(2 + i)
    style_header(ws[f"{col}23"])
    ws[f"{col}23"] = label
row = other_opex_start_row
for name, vals in other_opex_assumptions.items():
    txt(ws[f"B{row}"], name)
    for i, v in enumerate(vals):
        col = month_cols_salary[i]
        num(ws[f"{col}{row}"], v, fill=SOURCE_FILL)
    row += 1

ws["B31"] = "Распределение CAPEX по месяцам (для денежного потока)"
ws["B31"].font = BOLD
for i, m in enumerate(MONTHS):
    col = month_cols_salary[i]
    style_header(ws[f"{col}31"])
    ws[f"{col}31"] = m
for i, v in enumerate(capex_by_month):
    col = month_cols_salary[i]
    num(ws[f"{col}{capex_month_row}"], v, fill=SOURCE_FILL)

ws["B34"] = "CAPEX-исходники (для листа CAPEX)"
ws["B34"].font = BOLD
for c, label in zip("BCD", ["Статья", "CNY", "Комментарий"]):
    style_header(ws[f"{c}34"])
    ws[f"{c}34"] = label
row = capex_items_start_row
for name, cny, comment in capex_items:
    txt(ws[f"B{row}"], name)
    num(ws[f"C{row}"], cny, fill=SOURCE_FILL)
    txt(ws[f"D{row}"], comment)
    row += 1

# === OPEX помесячно ===
ws = wb.create_sheet("OPEX помесячно")
ws.sheet_view.showGridLines = False
for c, w in [("A", 4), ("B", 34), ("I", 14)]:
    ws.column_dimensions[c].width = w
for c in "CDEFGH":
    ws.column_dimensions[c].width = 12
ws["B2"] = "OPEX помесячно (CNY)"
ws["B2"].font = TITLE_FONT
for i, label in enumerate(["Статья"] + MONTHS + ["Итого"]):
    col = get_column_letter(2 + i)
    style_header(ws[f"{col}4"])
    ws[f"{col}4"] = label

rows_map = {
    5: "ФОТ (gross)",
    6: "Соцвзносы",
    7: "Аутсорс-бухгалтерия",
    8: "Аренда офиса",
    9: "Командировки и QC-выезды",
    10: "IT / связь (opex)",
    11: "Юр./комплаенс-ретейнер",
    12: "Прочее (перевод, ресёрч, hospit.)",
}
for r, name in rows_map.items():
    txt(ws[f"B{r}"], name)

for m_idx, col in enumerate("CDEFGH"):
    ws[f"{col}5"] = salary_formula(m_idx)
    num(ws[f"{col}5"], None)
    ws[f"{col}6"] = social_formula(m_idx)
    num(ws[f"{col}6"], None)

assumption_rows = {7: 24, 8: 25, 9: 26, 10: 27, 11: 28, 12: 29}
for target_row, src_row in assumption_rows.items():
    for m_idx, col in enumerate("CDEFGH"):
        src_col = month_cols_salary[m_idx]
        ws[f"{col}{target_row}"] = f"='{salary_sheet_name}'!{src_col}{src_row}"
        num(ws[f"{col}{target_row}"], None)

for r in range(5, 13):
    ws[f"I{r}"] = f"=SUM(C{r}:H{r})"
    num(ws[f"I{r}"], None)

txt(ws["B13"], "Подытог (до налога RO)", bold=True, fill=SUBTOTAL_FILL)
for col in "CDEFGH":
    ws[f"{col}13"] = f"=SUM({col}5:{col}12)"
    num(ws[f"{col}13"], None, bold=True, fill=SUBTOTAL_FILL)
ws["I13"] = "=SUM(C13:H13)"
num(ws["I13"], None, bold=True, fill=SUBTOTAL_FILL)

txt(ws["B14"], "Налог RO (deemed)")
for col in "CDEFGH":
    ws[f"{col}14"] = f"=ROUND({col}13*'{salary_sheet_name}'!${rate_tax_cell},0)"
    num(ws[f"{col}14"], None)
ws["I14"] = "=SUM(C14:H14)"
num(ws["I14"], None)

txt(ws["B15"], "ИТОГО OPEX/мес", bold=True, fill=TOTAL_FILL)
for col in "CDEFGH":
    ws[f"{col}15"] = f"={col}13+{col}14"
    num(ws[f"{col}15"], None, bold=True, fill=TOTAL_FILL)
ws["I15"] = "=SUM(C15:H15)"
num(ws["I15"], None, bold=True, fill=TOTAL_FILL)

txt(ws["B16"], "ИТОГО OPEX/мес, RUB")
for col in "CDEFGH":
    ws[f"{col}16"] = f"={col}15*'{salary_sheet_name}'!${rate_rub_cell}"
    num(ws[f"{col}16"], None)
ws["I16"] = "=SUM(C16:H16)"
num(ws["I16"], None)

# === Денежный поток ===
ws = wb.create_sheet("Денежный поток")
ws.sheet_view.showGridLines = False
for c, w in [("A", 4), ("B", 30), ("I", 14)]:
    ws.column_dimensions[c].width = w
for c in "CDEFGH":
    ws.column_dimensions[c].width = 12
ws["B2"] = "Денежный поток (CAPEX + OPEX), CNY"
ws["B2"].font = TITLE_FONT
for i, label in enumerate(["Статья"] + MONTHS + ["Итого"]):
    col = get_column_letter(2 + i)
    style_header(ws[f"{col}4"])
    ws[f"{col}4"] = label

txt(ws["B5"], "CAPEX", fill=CAPEX_FILL)
for m_idx, col in enumerate("CDEFGH"):
    src_col = month_cols_salary[m_idx]
    ws[f"{col}5"] = f"='{salary_sheet_name}'!{src_col}{capex_month_row}"
    num(ws[f"{col}5"], None, fill=CAPEX_FILL)
ws["I5"] = "=SUM(C5:H5)"
num(ws["I5"], None, fill=CAPEX_FILL)

txt(ws["B6"], "OPEX")
for col in "CDEFGH":
    ws[f"{col}6"] = f"='OPEX помесячно'!{col}15"
    num(ws[f"{col}6"], None)
ws["I6"] = "=SUM(C6:H6)"
num(ws["I6"], None)

txt(ws["B7"], "Итого/мес", bold=True, fill=TOTAL_FILL)
for col in "CDEFGH":
    ws[f"{col}7"] = f"={col}5+{col}6"
    num(ws[f"{col}7"], None, bold=True, fill=TOTAL_FILL)
ws["I7"] = "=SUM(C7:H7)"
num(ws["I7"], None, bold=True, fill=TOTAL_FILL)

txt(ws["B8"], "Итого/мес, RUB")
for col in "CDEFGH":
    ws[f"{col}8"] = f"={col}7*'{salary_sheet_name}'!${rate_rub_cell}"
    num(ws[f"{col}8"], None)
ws["I8"] = "=SUM(C8:H8)"
num(ws["I8"], None)

out = "Бюджет_представительство_КНР_6мес.xlsx"
wb.save(out)
print("Saved:", out)
